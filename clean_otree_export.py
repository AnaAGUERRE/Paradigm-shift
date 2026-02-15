# MAIN PURPOSE
    # This script processes and cleans data exported from oTree. 
    # It reads a CSV file (otree_exported_data.csv), 
    # restructures and filters the data, and outputs a clean_file.csv file 
    # with selected columns, renamed headers, and some data transformations.

# RELATIONS TO OTHER FILES
    # The script is designed to be run directly: python clean_otree_export.py
    #It is not imported or called by other files in the project.


import csv
import openpyxl  # For creating Excel files
from openpyxl.utils import get_column_letter  # For column width adjustment

# Input and output file names
INPUT_FILE = 'otree_exported_data.csv'
OUTPUT_FILE = 'clean_file.csv'

# Columns to extract and their new order/name
COLUMNS = [
    ('participant.code', 'Participant_code'),
    ('player.treatment', 'Treatment'),
    ('player.phase', 'Phase'),
    ('subsession.round_number', 'Round'),
    ('player.flower_colors', 'Flower_colors'),
    ('player.nutrient_choice', 'Nutrient_choice'),
    ('player.score_per_flower', 'Score_per_flower'),
    ('player.Vround', 'Vround'),
    ('player.score_reel', 'Score_reel_flower'),

    ('player.cumulative_earnings', 'Total_earnings'),
    ('player.popup_strategy', 'Popup_strategy'),
    ('player.year_of_birth', 'Birth_year'),
    ('player.feedback', 'Feedback'),
    ('player.qcm_click_sequence', 'QCM_sequence_raw'),
    ('QCM_sequence_bool', 'QCM_sequence_bool'),
]

def clean_otree_export(input_path=INPUT_FILE, output_path=OUTPUT_FILE):
    # Open the input CSV file
    with open(input_path, newline='', encoding='utf-8') as infile:
        reader = csv.DictReader(infile, delimiter=';')
        rows = list(reader)

    # Group rows by participant code
    from collections import defaultdict
    by_participant = defaultdict(list)
    for row in rows:
        code = row.get('participant.code', '')
        by_participant[code].append(row)

    output_rows = []
    import json
    for code, rounds in by_participant.items():
        # Find the last round for each participant
        last_round = None
        max_round = -1
        for row in rounds:
            try:
                r = int(row.get('subsession.round_number', '0'))
            except Exception:
                r = 0
            if r > max_round:
                max_round = r
                last_round = row

        # For each row, process and clean data
        for row in rounds:
            output_row = {}
            phase = row.get('player.phase', '')
            round_num = row.get('subsession.round_number', '')
            # If this is the last round, set phase to 'Test 2'
            if round_num and int(round_num) == max_round:
                phase = 'Test 2'
            # Only keep feedback/birthyear for summary row
            for old, new in COLUMNS:
                if new in ('Feedback', 'Birth_year'):
                    output_row[new] = ''
                elif new == 'QCM_sequence_bool':
                    # Parse QCM sequence JSON and extract
                    seq_raw = row.get('player.qcm_click_sequence', '')
                    try:
                        seq = json.loads(seq_raw) if seq_raw else []
                        output_row[new] = str([item.get('correct', False) for item in seq])
                    except Exception:
                        output_row[new] = ''
                else:
                    output_row[new] = row.get(old, '')
            output_row['Phase'] = phase
            output_row['Round'] = round_num
            output_rows.append(output_row)

        # Add a summary row for each participant
        summary = {new: '' for _, new in COLUMNS}
        if last_round:
            summary['Participant_code'] = last_round.get('participant.code', '')
            summary['Treatment'] = last_round.get('player.treatment', '')
            summary['Phase'] = 'Results'
            summary['Round'] = ''
            summary['Total_earnings'] = last_round.get('player.cumulative_earnings', '')
            summary['Birth_year'] = last_round.get('player.year_of_birth', '')
            summary['Feedback'] = last_round.get('player.feedback', '')
            # All other fields remain blank
        output_rows.append(summary)


    # Write the clean data to an Excel file with auto-fit columns
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    # Write header
    ws.append([col[1] for col in COLUMNS])
    # Write data rows
    for row in output_rows:
        ws.append([row[col[1]] for col in COLUMNS])
    # Auto-fit columns
    for col_idx, col in enumerate(COLUMNS, 1):
        max_length = len(col[1])
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in row:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value is not None else 0)
                except:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    wb.save(output_path.replace('.csv', '.xlsx'))

# This ensures the script only processes the data when run directly,
# not when imported elsewhere.
if __name__ == '__main__':
    clean_otree_export() 

"""
Flower Field Task — version indépendante (pas de transmission)
Main oTree app logic for the Flower Field experiment.
Handles session setup, round logic, earnings, and data export.
"""
doc = """Flower Field Task — version indépendante (pas de transmission)"""


# oTree imports and engine functions
from otree.api import *
from .engine import run_engine, calculate_points_from_growth
import json

class C(BaseConstants):
    NAME_IN_URL = 'flowerfieldtask'  # URL name for this app
    PLAYERS_PER_GROUP = None         # No grouping
    TRAINING_ROUNDS = 4              # Number of training rounds
    TEST1_ROUNDS = 1                 # Number of test1 rounds
    EXPLORATION_ROUNDS = 5           # Number of exploration rounds
    TEST2_ROUNDS = 1                 # Number of test2 rounds
    NUM_ROUNDS = TRAINING_ROUNDS + TEST1_ROUNDS + EXPLORATION_ROUNDS + TEST2_ROUNDS  # Total rounds

class Subsession(BaseSubsession):
    def creating_session(self):
        # Initialize total earnings for each participant
        for p in self.session.get_participants():
            if 'total_earnings' not in p.vars:
                p.vars['total_earnings'] = 0
        # Set cumulative earnings for each player
        for player in self.get_players():
            player.cumulative_earnings = player.participant.vars.get('total_earnings', 0)

class Group(BaseGroup):
    pass  # No group logic needed

class Player(BasePlayer):
    cumulative_earnings = models.FloatField(initial=0)  # Tracks total earnings for this player
    test1_pending_earnings = models.FloatField(initial=0)  # Holds Test 1 earnings until Test 2


class Instructions(Page):
    def is_displayed(player):
        return player.round_number == 1
    def vars_for_template(player):
        return {}
    template_name = '_templates/instructions.html'

class FlowerField(Page):
    import json
    def vars_for_template(player):
        # Pass variables to the template for display and JS logic
        # For Test 1, show total_earnings without Test 1 reward
        # For Test 2, show total_earnings including both Test 1 and Test 2 rewards
        phase = None
        if player.round_number <= C.TRAINING_ROUNDS:
            phase = 'Training phase'
        elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS:
            phase = 'Test 1'
        elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS + C.EXPLORATION_ROUNDS:
            phase = 'Exploration phase'
        else:
            phase = 'Test 2'
        if phase == 'Test 1':
            player.cumulative_earnings = player.participant.vars.get('total_earnings', 0)
        else:
            player.cumulative_earnings = player.participant.vars.get('total_earnings', 0)
        # Determine phase and flower colors for this round
        treatment = player.session.config.get('display_name', '').lower()
        # Map noisy and transmission treatments to their no-noise logic for flower sequences
        if treatment == 'anomaly noisy':
            treatment_logic = 'anomaly no noise'
        elif treatment == 'no anomaly noisy':
            treatment_logic = 'no anomaly no noise'
        elif treatment in ['transmission correct', 'transmission m&m']:
            treatment_logic = 'anomaly no noise'
        else:
            treatment_logic = treatment
        if player.round_number <= C.TRAINING_ROUNDS:
            phase = 'Training phase'
            phase_round = player.round_number
            phase_total = C.TRAINING_ROUNDS
            round_flower_types = [
                ['Purple', 'Green'],      # R1
                ['Green', 'Purple'],      # R2
                ['Purple', 'Green', 'Green'], # R3
                ['Green', 'Purple', 'Purple'] # R4
            ]
            flower_colors = round_flower_types[player.round_number - 1]
        elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS:
            phase = 'Test 1'
            phase_round = player.round_number - C.TRAINING_ROUNDS
            phase_total = C.TEST1_ROUNDS
            flower_colors = ['Green', 'Yellow', 'Purple', 'Red', 'Orange', 'Blue']
        elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS + C.EXPLORATION_ROUNDS:
            phase = 'Exploration phase'
            phase_round = player.round_number - C.TRAINING_ROUNDS - C.TEST1_ROUNDS
            phase_total = C.EXPLORATION_ROUNDS
            if treatment_logic == 'anomaly no noise':
                exploration_flower_types = [
                    ['Orange', 'Purple'],         # R1
                    ['Orange', 'Green'],          # R2
                    ['Green', 'Red', 'Green'],    # R3
                    ['Blue', 'Purple', 'Purple'], # R4
                    ['Purple', 'Green', 'Yellow'] # R5
                ]
            elif treatment_logic == 'no anomaly no noise':
                exploration_flower_types = [
                    ['Orange', 'Orange'],         # R1
                    ['Purple', 'Green'],          # R2
                    ['Red', 'Blue', 'Yellow'],    # R3
                    ['Green', 'Purple', 'Green'], # R4
                    ['Purple', 'Purple', 'Green'] # R5
                ]
            else:
                exploration_flower_types = [
                    ['Green', 'Purple', 'Blue'],
                    ['Green', 'Purple', 'Blue'],
                    ['Green', 'Purple', 'Yellow'],
                    ['Green', 'Purple', 'Yellow'],
                    ['Green', 'Purple', 'Yellow']
                ]
            flower_colors = exploration_flower_types[phase_round - 1]
        else:
            phase = 'Test 2'
            phase_round = player.round_number - C.TRAINING_ROUNDS - C.TEST1_ROUNDS - C.EXPLORATION_ROUNDS
            phase_total = C.TEST2_ROUNDS
            flower_colors = ['Green', 'Yellow', 'Purple', 'Red', 'Orange', 'Blue']
        # Set phase_class for template
        if phase == 'Training phase':
            phase_class = 'training-phase-flowers'
        elif phase == 'Test 1':
            phase_class = 'test-1-flowers'
        elif phase == 'Exploration phase':
            phase_class = 'exploration-flowers'
        elif phase == 'Test 2':
            phase_class = 'test-1-flowers'
        else:
            phase_class = ''
        # Prepare previous rounds' history for Training and Exploration phases
        previous_combinations = []
        valid_flowers = ['Blue', 'Green', 'Orange', 'Purple', 'Red', 'Yellow']
        if phase in ["Training phase", "Exploration phase"]:
            history = player.participant.vars.get('nutrient_flower_history', [])
            # Only show previous rounds (not current), and only those matching the current round_flower_types
            round_flower_types = [
                ['Purple', 'Green'],
                ['Green', 'Purple'],
                ['Purple', 'Green', 'Green'],
                ['Green', 'Purple', 'Purple']
            ]
            for entry in history:
                if entry['phase'] == phase and entry['round'] < phase_round:
                    # Always use the expected flower set for this round
                    expected_flowers = round_flower_types[entry['round'] - 1] if phase == 'Training phase' and 1 <= entry['round'] <= 4 else entry['flower_colors']
                    # Use growth values for flower size, not pennies
                    growths = entry.get('growths', None)
                    if not growths:
                        growths = []
                        for s in entry.get('scores', []):
                            try:
                                growths.append(float(s) / 10.0)
                            except:
                                growths.append(0.0)
                    filtered = [
                        (
                            f,
                            [f"img/Nutr{nut}.png" if nut else None for nut in (n if n is not None else [])],
                            (s if s is not None else '0'),  # per-flower earnings as string (pennies, show 0 if missing)
                            f"img/Flw{f}.png" if f is not None and f != '' and f in valid_flowers else None,
                            g,  # growth value as float (for proportional size)
                            int(18 + (g if g is not None else 0.0) * 18)  # flower_size in px (smaller base and scale)
                        )
                        for (f, n, s, g) in zip(expected_flowers, entry.get('nutrients', []), entry.get('scores', []), growths)
                        if f is not None and f != '' and f in valid_flowers
                    ]
                    previous_combinations.append({
                        'round': entry['round'],
                        'zipped': filtered
                    })
        valid_flowers = ['Purple', 'Orange', 'Green', 'Yellow', 'Red', 'Blue']
        # Centering logic for flower placement (pass count to template)
        flower_count = len(flower_colors)
        # Determine treatment for template logic
        treatment = player.session.config.get('display_name', '')
        # Example: transmitted_photo could be set in participant.vars or elsewhere
        transmitted_photo = player.participant.vars.get('transmitted_photo', None)
        return dict(
            phase=phase,
            phase_round=phase_round,
            phase_total=phase_total,
            flower_colors=json.dumps(flower_colors),
            cumulative_earnings=player.cumulative_earnings,
            phase_class=phase_class,
            previous_combinations=previous_combinations,
            valid_flowers=valid_flowers,
            flower_count=flower_count,
            treatment=treatment,
            transmitted_photo=transmitted_photo
        )
    live_method = "live_method"  # Name of live method for JS communication
    template_name = 'flowerfieldtask/FlowerField.html'  # HTML template to use


    def live_method(player, data):
        # Handles live communication from JS (nutrient submission)
        if data['type'] == 'flowerSubmit':
            nutrients = data['data']
            # Select scoring system from session config
            scoring_system = player.session.config.get('scoring_system', 'anomaly')
            # Use the same treatment-dependent logic as vars_for_template
            treatment = player.session.config.get('display_name', '').lower()
            # Map noisy and transmission treatments to their no-noise logic for flower sequences
            if treatment == 'anomaly noisy':
                treatment_logic = 'anomaly no noise'
            elif treatment == 'no anomaly noisy':
                treatment_logic = 'no anomaly no noise'
            elif treatment in ['transmission correct', 'transmission m&m']:
                treatment_logic = 'anomaly no noise'
            else:
                treatment_logic = treatment
            if player.round_number <= C.TRAINING_ROUNDS:
                phase = 'Training phase'
                display_round = player.round_number
                round_flower_types = [
                    ['Purple', 'Green'],      # R1
                    ['Green', 'Purple'],      # R2
                    ['Purple', 'Green', 'Green'], # R3
                    ['Green', 'Purple', 'Purple'] # R4
                ]
                flower_colors = round_flower_types[player.round_number - 1]
            elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS:
                phase = 'Test 1'
                display_round = player.round_number - C.TRAINING_ROUNDS
                flower_colors = ['Green', 'Yellow', 'Purple', 'Red', 'Orange', 'Blue']
            elif player.round_number <= C.TRAINING_ROUNDS + C.TEST1_ROUNDS + C.EXPLORATION_ROUNDS:
                phase = 'Exploration phase'
                display_round = player.round_number - C.TRAINING_ROUNDS - C.TEST1_ROUNDS
                if treatment_logic == 'anomaly no noise':
                    exploration_flower_types = [
                        ['Orange', 'Purple'],         # R1
                        ['Orange', 'Green'],          # R2
                        ['Green', 'Red', 'Green'],    # R3
                        ['Blue', 'Purple', 'Purple'], # R4
                        ['Purple', 'Green', 'Yellow'] # R5
                    ]
                elif treatment_logic == 'no anomaly no noise':
                    exploration_flower_types = [
                        ['Orange', 'Orange'],         # R1
                        ['Purple', 'Green'],          # R2
                        ['Red', 'Blue', 'Yellow'],    # R3
                        ['Green', 'Purple', 'Green'], # R4
                        ['Purple', 'Purple', 'Green'] # R5
                    ]
                else:
                    exploration_flower_types = [
                        ['Green', 'Purple', 'Blue'],
                        ['Green', 'Purple', 'Blue'],
                        ['Green', 'Purple', 'Yellow'],
                        ['Green', 'Purple', 'Yellow'],
                        ['Green', 'Purple', 'Yellow']
                    ]
                flower_colors = exploration_flower_types[display_round - 1]
            else:
                phase = 'Test 2'
                display_round = player.round_number - C.TRAINING_ROUNDS - C.TEST1_ROUNDS - C.EXPLORATION_ROUNDS
                flower_colors = ['Green', 'Yellow', 'Purple', 'Red', 'Orange', 'Blue']
            # Run backend engine to calculate growth and points
            if scoring_system == 'mm':
                output = run_engine(nutrients, flower_colors=flower_colors, scoring_system='mm')
            else:
                output = run_engine(nutrients)
            total_growth = sum(f['growth'] for f in output) / len(output)
            total_points = calculate_points_from_growth(total_growth)
            flower_scores = [f['growth'] for f in output]
            # Multiply by 2 only for Test 1 and Test 2
            if phase in ['Test 1', 'Test 2']:
                flower_scores = [g * 2 for g in flower_scores]
            # Each earning: growth * 10 (to get pennies), rounded to nearest int
            flower_earnings = [str(int(round(g * 10))) for g in flower_scores]
            # For total, sum as £
            round_earnings = sum([int(e) for e in flower_earnings]) / 100.0
            # Track noise effects for noisy configs
            noise_effects = [f.get('noise') for f in output]
            # Update participant's total earnings
            if 'total_earnings' not in player.participant.vars:
                player.participant.vars['total_earnings'] = 0.0
            if phase == 'Test 1':
                # Store Test 1 earnings but do not add to total yet
                player.test1_pending_earnings = round_earnings
                player.participant.vars['test1_pending_earnings'] = round_earnings
                # Do not update total_earnings yet
            elif phase == 'Test 2':
                # Add both Test 1 and Test 2 earnings to total_earnings
                test1_earnings = player.participant.vars.get('test1_pending_earnings', 0.0)
                player.participant.vars['total_earnings'] = round(float(player.participant.vars['total_earnings']) + test1_earnings + round_earnings, 2)
                player.cumulative_earnings = player.participant.vars['total_earnings']
                player.participant.vars['test1_pending_earnings'] = 0.0
            else:
                # For all other phases, update as usual
                player.participant.vars['total_earnings'] = round(float(player.participant.vars['total_earnings']) + round_earnings, 2)
                player.cumulative_earnings = player.participant.vars['total_earnings']

            # (Removed duplicate/old block: always use the flower_colors and phase already set above)

            # Store nutrient-flower combinations and flower colors for each round
            if 'nutrient_flower_history' not in player.participant.vars:
                player.participant.vars['nutrient_flower_history'] = []
            # Ensure all lists are the same length and order as flower_colors
            n_flowers = len(flower_colors)
            safe_nutrients = (nutrients if len(nutrients) == n_flowers else [None]*n_flowers)
            safe_scores = (flower_earnings if len(flower_earnings) == n_flowers else [None]*n_flowers)
            safe_growths = (flower_scores if len(flower_scores) == n_flowers else [None]*n_flowers)
            safe_noise = (noise_effects if len(noise_effects) == n_flowers else [None]*n_flowers)
            player.participant.vars['nutrient_flower_history'].append({
                'phase': phase,
                'round': display_round,
                'flower_colors': list(flower_colors),
                'nutrients': list(safe_nutrients),
                'scores': list(safe_scores),
                'growths': list(safe_growths),
                'noise_effects': list(safe_noise)
            })

            # Save all entries to Excel after every round (local analysis)
            try:
                import pandas as pd
                import os
                # Prepare the new entry for Excel
                new_entry = {
                    'participant_code': player.participant.code,
                    'Treatment': player.session.config.get('display_name', player.session.config.get('name', '')),
                    'phase': phase,
                    'round': display_round,
                    'flower_colors': flower_colors,
                    'nutrients': nutrients,
                    'scores (p)': flower_earnings,  # in pennies as strings
                    'noise_effects': noise_effects
                }
                excel_path = 'nutrient_flower_data.xlsx'
                if os.path.exists(excel_path):
                    df_existing = pd.read_excel(excel_path)
                    df_new = pd.DataFrame([new_entry])
                    df_all = pd.concat([df_existing, df_new], ignore_index=True)
                else:
                    df_all = pd.DataFrame([new_entry])

                # Remove any old 'final_total_earnings' or 'final_total_earnings (£)' column if present
                for col in ['final_total_earnings', 'final_total_earnings (£)']:
                    if col in df_all.columns:
                        df_all = df_all.drop(columns=[col])
                # Only keep 'final_total_earnings (£)' and only fill it for the last row of each participant
                df_all['final_total_earnings (£)'] = None
                for code in df_all['participant_code'].unique():
                    mask = df_all['participant_code'] == code
                    # Only include non-test phases
                    mask_non_test = mask & (~df_all['phase'].isin(['Test 1', 'Test 2']))
                    total = 0.0
                    for idx, row in df_all[mask_non_test].iterrows():
                        # Use the 'scores (p)' column, which is a list of penny strings
                        scores_p = row.get('scores (p)')
                        if isinstance(scores_p, list):
                            try:
                                total += sum([int(s) for s in scores_p])
                            except:
                                pass
                        elif isinstance(scores_p, str):
                            try:
                                import ast
                                scores = ast.literal_eval(scores_p)
                                total += sum([int(s) for s in scores])
                            except:
                                pass
                    last_idx = df_all[mask].index[-1]
                    # Convert pennies to £ and format to two decimals
                    df_all.at[last_idx, 'final_total_earnings (£)'] = '{:.2f}'.format(total / 100)

                # Reorder columns to put 'Treatment' first
                cols = list(df_all.columns)
                if 'Treatment' in cols:
                    cols.insert(0, cols.pop(cols.index('Treatment')))
                    df_all = df_all[cols]
                df_all.to_excel(excel_path, index=False)

                # Auto-adjust column widths for readability
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            cell_length = len(str(cell.value)) if cell.value is not None else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                wb.save(excel_path)
            except Exception as e:
                print(f"Excel export error: {e}")

            # Return results to JS for display
            # Only after Test 2 submission, send both Test 1 and Test 2 raw data for animation
            test1_data = None
            test2_data = None
            if phase == 'Test 2':
                history = player.participant.vars.get('nutrient_flower_history', [])
                # Find the first Test 1 and the last Test 2 entries in history
                test1_entry = None
                test2_entry = None
                for entry in history:
                    if entry.get('phase') == 'Test 1' and test1_entry is None:
                        test1_entry = entry
                for entry in reversed(history):
                    if entry.get('phase') == 'Test 2' and test2_entry is None:
                        test2_entry = entry
                        break
                test1_data = dict(
                    flower_colors = test1_entry['flower_colors'] if test1_entry else [],
                    nutrients = test1_entry['nutrients'] if test1_entry else [],
                    scores = test1_entry['scores'] if test1_entry else [],
                    growths = test1_entry['growths'] if test1_entry else []
                )
                test2_data = dict(
                    flower_colors = test2_entry['flower_colors'] if test2_entry else [],
                    nutrients = test2_entry['nutrients'] if test2_entry else [],
                    scores = test2_entry['scores'] if test2_entry else [],
                    growths = test2_entry['growths'] if test2_entry else []
                )
                print("[FEEDBACK] Test 1 and Test 2 results found and sent for animation.")
            result_dict = dict(
                flower_scores=flower_scores,
                flower_earnings=flower_earnings,
                cumulative_earnings="{:.2f}".format(player.cumulative_earnings)
            )
            if phase == 'Test 2':
                result_dict['test1_data'] = test1_data
                result_dict['test2_data'] = test2_data
            return {player.id_in_group: result_dict}



# Survey page at the end
class Survey(Page):

        form_model = 'player'
        form_fields = ['year_of_birth', 'feedback']

        def before_next_page(player, timeout_happened):
            # On survey page, add a single row with phase='Survey' for this participant
            import pandas as pd
            import os
            excel_path = 'nutrient_flower_data.xlsx'
            if os.path.exists(excel_path):
                df_all = pd.read_excel(excel_path)
            else:
                df_all = pd.DataFrame()
            # Add a new row for the survey
            new_entry = {
                'participant_code': player.participant.code,
                'phase': 'Survey',
                'year_of_birth': player.year_of_birth,
                'feedback': player.feedback,
            }
            df_all = pd.concat([df_all, pd.DataFrame([new_entry])], ignore_index=True)
            df_all.to_excel(excel_path, index=False)
            # Auto-adjust column widths for readability
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                ws = wb.active
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try:
                            cell_length = len(str(cell.value)) if cell.value is not None else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                wb.save(excel_path)
            except Exception as e:
                print(f"Excel auto-width error: {e}")

        def vars_for_template(player):
            # Years from 1940 to current year
            import datetime
            current_year = datetime.datetime.now().year
            years = list(range(current_year, 1939, -1))
            return dict(years=years)

        def is_displayed(player):
            # Only show after the last round (after Test2)
            return player.round_number == C.NUM_ROUNDS

        template_name = 'flowerfieldtask/survey.html'

# Add fields to Player model
Player.year_of_birth = models.IntegerField(blank=True, null=True, label="What is your year of birth?")
Player.feedback = models.LongStringField(blank=True, null=True, label="Do you have any feedback for us? Were there any problems? (optional)")


# Results page at the very end

# Results page at the very end
class Results(Page):
    def vars_for_template(player):
        # Use the participant's total earnings (in £, formatted)
        total = player.participant.vars.get('total_earnings', 0)
        # Always provide test1_zipped and test2_zipped as empty lists for template robustness
        return dict(
            total_earnings="{:.2f}".format(total),
            test1_zipped=[],
            test2_zipped=[],
            test1_result=None,
            test2_result=None
        )
    def is_displayed(player):
        return player.round_number == C.NUM_ROUNDS
    template_name = 'flowerfieldtask/results.html'

# TestResults page to show both Test 1 and Test 2 results after Test 2
class TestResults(Page):
    def is_displayed(player):
        # Only show after Test 2 round
        return player.round_number == C.TRAINING_ROUNDS + C.TEST1_ROUNDS + C.EXPLORATION_ROUNDS + C.TEST2_ROUNDS
    def vars_for_template(player):
        # Get test1_data and test2_data from participant vars (populated after Test 2)
        history = player.participant.vars.get('nutrient_flower_history', [])
        test1_entry = None
        test2_entry = None
        for entry in history:
            if entry.get('phase') == 'Test 1' and test1_entry is None:
                test1_entry = entry
        for entry in reversed(history):
            if entry.get('phase') == 'Test 2' and test2_entry is None:
                test2_entry = entry
                break
        test1_data = dict(
            flower_colors = test1_entry['flower_colors'] if test1_entry else [],
            nutrients = test1_entry['nutrients'] if test1_entry else [],
            scores = test1_entry['scores'] if test1_entry else [],
            growths = test1_entry['growths'] if test1_entry else []
        )
        test2_data = dict(
            flower_colors = test2_entry['flower_colors'] if test2_entry else [],
            nutrients = test2_entry['nutrients'] if test2_entry else [],
            scores = test2_entry['scores'] if test2_entry else [],
            growths = test2_entry['growths'] if test2_entry else []
        )
        import json
        # Add phase and earnings for top bar
        total = player.participant.vars.get('total_earnings', 0)
        return dict(
            test1_data=json.dumps(test1_data),
            test2_data=json.dumps(test2_data),
            phase='Test Results',
            phase_round='',
            phase_total='',
            cumulative_earnings="{:.2f}".format(total)
        )
    template_name = 'flowerfieldtask/test_results.html'

# Sequence of pages in the experiment
page_sequence = [Instructions, FlowerField, TestResults, Survey, Results]

